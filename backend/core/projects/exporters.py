# application/exporters.py

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from .models import Project
from materials.models import Category, Material
import math
import os

class ExcelExporter:
    """
    Génère un rapport de budget de projet au format Excel (.xlsx) avec design professionnel Bouygues.
    Utilise les vraies catégories et matériaux de la base de données.
    """

    def __init__(self):
        # Définition des styles pour le fichier Excel
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Bouygues Blue
        self.header_align = Alignment(horizontal="center", vertical="center")
        self.category_font = Font(bold=True, color="000000")
        self.category_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
        self.currency_format_eur = '"€"#,##0.00'
        self.currency_format_mad = '"MAD"#,##0.00'
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def generate_project_excel(self, project: Project):
        """
        Generate Excel file for project budget using REAL categories and materials from database.
        Structure exacte comme l'image de référence avec vraies données.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Budget Projet"

        # --- Company Logo (starting from line 3) ---
        # Try to add actual logo image if it exists
        logo_path = os.path.join(os.path.dirname(__file__), '..', '..', '..', 'frontend', 'public', 'logo.png')
        if os.path.exists(logo_path):
            try:
                logo = Image(logo_path)
                logo.width = 300  # Smaller width
                logo.height = 100  # Smaller height
                ws.add_image(logo, 'A3')
            except Exception as e:
                print(f"Could not load logo image: {e}")
                # Fallback to text logo
                ws.merge_cells('A1:C1')
                ws['A1'] = "🏗️ BOUYGUES CONSTRUCTION"
                ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
                ws['A1'].fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
                ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        else:
            # Create a professional text-based logo with better styling
            ws.merge_cells('A1:C1')
            ws['A1'] = "🏗️ BOUYGUES CONSTRUCTION"
            ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            
            # Add a professional border around the logo
            ws['A1'].border = Border(
                left=Side(style='thick', color='FFFFFF'),
                right=Side(style='thick', color='FFFFFF'),
                top=Side(style='thick', color='FFFFFF'),
                bottom=Side(style='thick', color='FFFFFF')
            )
        
        # --- Project Title (moved to line 14) ---
        ws.merge_cells('A14:I14')
        ws['A14'] = f"{project.name}: Estimate financial study"
        ws['A14'].font = Font(bold=True, size=16, color="1E3A8A")
        ws['A14'].alignment = Alignment(horizontal="center", vertical="center")
        
        # --- Project Info (centered in the middle) ---
        ws.merge_cells('C3:E3')  # Center the user count in the middle
        ws['C3'] = f"NB of users: {project.number_of_users}"
        ws['C3'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['C3'].fill = PatternFill(start_color="00AA00", end_color="00AA00", fill_type="solid")
        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        
        # --- Estimated Cost Section (centered in the middle) ---
        ws.merge_cells('F3:H3')  # Center the estimated cost header in the middle
        ws['F3'] = "Estimated cost (€)"
        ws['F3'].font = Font(bold=True, size=12, color="2D3748")
        ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Add totals with formulas (will be calculated after materials are added)
        # Add labels in column E and formulas in column F
        ws['E4'] = "One off:"
        ws['E4'].font = Font(bold=True, size=12, color="2D3748")
        ws['E4'].alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells('F4:H4')  # Center One off formula
        ws['F4'] = 0  # Will be updated with formula later
        ws['F4'].font = Font(bold=True, size=12, color="2D3748")
        ws['F4'].alignment = Alignment(horizontal="center", vertical="center")
        
        ws['E5'] = "Monthly:"
        ws['E5'].font = Font(bold=True, size=12, color="2D3748")
        ws['E5'].alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells('F5:H5')  # Center Monthly formula
        ws['F5'] = 0  # Will be updated with formula later
        ws['F5'].font = Font(bold=True, size=12, color="2D3748")
        ws['F5'].alignment = Alignment(horizontal="center", vertical="center")
        
        ws['E6'] = "Yearly:"
        ws['E6'].font = Font(bold=True, size=12, color="2D3748")
        ws['E6'].alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells('F6:H6')  # Center Yearly formula
        ws['F6'] = 0  # Will be updated with formula later
        ws['F6'].font = Font(bold=True, size=12, color="2D3748")
        ws['F6'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Add MONTHLY TOTAL (One off + Monthly) with green color
        ws['E7'] = "MONTHLY TOTAL:"
        ws['E7'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['E7'].fill = PatternFill(start_color="00AA00", end_color="00AA00", fill_type="solid")
        ws['E7'].alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells('F7:H7')  # Center the monthly total formula
        ws['F7'] = 0  # Will be updated with formula later
        ws['F7'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['F7'].fill = PatternFill(start_color="00AA00", end_color="00AA00", fill_type="solid")
        ws['F7'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Add YEARLY TOTAL (One off + Yearly) with green color
        ws['E8'] = "YEARLY TOTAL:"
        ws['E8'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['E8'].fill = PatternFill(start_color="00AA00", end_color="00AA00", fill_type="solid")
        ws['E8'].alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells('F8:H8')  # Center the yearly total formula
        ws['F8'] = 0  # Will be updated with formula later
        ws['F8'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['F8'].fill = PatternFill(start_color="00AA00", end_color="00AA00", fill_type="solid")
        ws['F8'].alignment = Alignment(horizontal="center", vertical="center")
        
        # --- Column Headers (matching the image structure) ---
        # Add "France purchase" header spanning all cost columns with blue background
        ws.merge_cells('E20:J20')
        ws['E20'] = "France purchase"
        ws['E20'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['E20'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        ws['E20'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E20'].border = self.thin_border
        
        # Main headers (row 21) - materials start in column B, not A
        ws.cell(row=21, column=1, value="")  # Empty for categories
        ws.cell(row=21, column=2, value="Item")
        ws.cell(row=21, column=3, value="Quantity")
        ws.cell(row=21, column=4, value="Description")
        
        # Cost header (merged E21-F21)
        ws.merge_cells('E21:F21')
        ws['E21'] = "Cost"
        ws['E21'].font = Font(bold=True, color="FFFFFF")
        ws['E21'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        ws['E21'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E21'].border = self.thin_border
        
        # Monthly cost header (merged G21-H21)
        ws.merge_cells('G21:H21')
        ws['G21'] = "Monthly cost"
        ws['G21'].font = Font(bold=True, color="FFFFFF")
        ws['G21'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        ws['G21'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G21'].border = self.thin_border
        
        # Yearly cost header (merged I21-J21) - now only 2 columns instead of 3
        ws.merge_cells('I21:J21')
        ws['I21'] = "Yearly cost"
        ws['I21'].font = Font(bold=True, color="FFFFFF")
        ws['I21'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        ws['I21'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I21'].border = self.thin_border
        
        # Sub-headers (row 22)
        ws.cell(row=22, column=1, value="")  # Empty for categories
        ws.cell(row=22, column=2, value="")  # Empty for Item
        ws.cell(row=22, column=3, value="")  # Empty for Quantity
        ws.cell(row=22, column=4, value="")  # Empty for Description
        
        ws.cell(row=22, column=5, value="Unit. €")
        ws.cell(row=22, column=6, value="Total €")
        ws.cell(row=22, column=7, value="Unit. €")
        ws.cell(row=22, column=8, value="Total €")
        # Yearly cost Total € spans 2 columns (I22-J22)
        ws.merge_cells('I22:J22')
        ws['I22'] = "Total €"
        
        # Style the sub-header row (columns E to J)
        for col in range(5, 11):
            cell = ws.cell(row=22, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.thin_border
        
        # Style the main headers (columns B to D)
        for col in [2, 3, 4]:
            cell = ws.cell(row=21, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.thin_border

        # --- Category and Item Data ---
        row_cursor = 23  # Data starts from row 23 (after headers)
        
        # Get REAL categories from database
        categories = Category.objects.all().order_by('name')
        
        for category in categories:
            # Add category header row (yellow background) - ONLY in column A
            ws.cell(row=row_cursor, column=1, value=category.name)
            cell = ws.cell(row=row_cursor, column=1)
            cell.fill = self.category_fill
            cell.font = self.category_font
            cell.border = self.thin_border
            row_cursor += 1

            # Add EMPTY ROW for spacing (materials will start in next row, column B)
            ws.append(["", "", "", "", "", "", "", ""])
            row_cursor += 1

            # Get REAL materials for this category
            materials = Material.objects.filter(category=category)
            
            # Get project items for this category using the new ProjectItem model
            from calculations.models import ProjectItem
            project_items = ProjectItem.objects.filter(
                project=project,
                material__category=category
            ).select_related('material')

            # Get ALL quantities for this project using the REAL calculator
            from calculations.services import ProjectCalculator
            calculator = ProjectCalculator()
            
            # Get all types of items
            user_items = calculator.get_user_specified_items(project)
            auto_items = calculator.calculate_automatic_items(project)
            service_items = calculator.get_service_items(project)
            custom_items = calculator.calculate_custom_materials(project)
            
            # Combine all items, avoiding duplicates (same logic as in calculate_budget)
            all_items = {}
            all_items.update(user_items)  # User-specified items first (highest priority)
            
            # Add auto-calculated items (only if not already in user items)
            for item_name, quantity in auto_items.items():
                if item_name not in all_items:
                    all_items[item_name] = quantity
            
            # Add service items (only if not already calculated)
            for item_name, quantity in service_items.items():
                if item_name not in all_items:
                    all_items[item_name] = quantity
            
            # Add custom items (only if not already calculated)
            for item_name, quantity in custom_items.items():
                if item_name not in all_items:
                    all_items[item_name] = quantity
            
            # Filter internet services - only show the selected one
            selected_internet_service = None
            if project.internet_line_type and project.internet_line_speed:
                # Map internet line types to their full names
                type_mapping = {
                    'FO': 'Fiber Optic',
                    'STARLINK': 'STARLINK',
                    'VSAT': 'VSAT'
                }
                full_type = type_mapping.get(project.internet_line_type, project.internet_line_type)
                selected_internet_service = f"{full_type} {project.internet_line_speed}"
            
            # Add materials under this category - prioritize materials with quantities
            materials_with_qty = []
            materials_without_qty = []
            
            for material in materials:
                # Skip unselected internet services
                if self._is_internet_service(material.name):
                    if selected_internet_service and selected_internet_service not in material.name:
                        continue  # Skip this internet service as it's not selected
                # First check if this material has calculated quantity
                qty = all_items.get(material.name, 0)
                
                # Also check project items for additional quantities
                project_item_qty = 0
                for item in project_items:
                    if item.material and item.material.name == material.name:
                        project_item_qty += item.quantity
                        break
                
                # Use the maximum of auto-calculated or project item quantity
                final_qty = max(qty, project_item_qty)
                
                
                if final_qty > 0:
                    materials_with_qty.append((material, final_qty))
                else:
                    materials_without_qty.append((material, final_qty))
            
            # First show materials with quantities, then materials without quantities
            all_materials = materials_with_qty + materials_without_qty
            
            for material, qty in all_materials:
                # Show ALL materials, even with 0 quantity
                
                # Get material details
                name = material.name
                description = material.description or ""
                unit_cost = float(material.price_france or 0)
                
                # Monthly cost logic: Only for specific categories
                monthly_cost = 0
                if category.name in ['BYCN IT costs', 'Services', 'Software Licenses']:
                    # For these categories, use the same price as unit cost for monthly
                    monthly_cost = unit_cost
                    # Set unit_cost to 0 for these categories (they only have monthly costs)
                    unit_cost = 0
                
                # Yearly cost logic: For ALL materials, yearly = monthly * 12
                yearly_cost = monthly_cost * 12
                
                unit = material.unit or "unit"
                
                # Create row with materials starting in column B (not A) - categories stay in A
                ws.cell(row=row_cursor, column=2, value=name)  # Item in column B
                ws.cell(row=row_cursor, column=3, value=qty)  # Quantity in column C
                ws.cell(row=row_cursor, column=4, value=description)  # Description in column D
                
                ws.cell(row=row_cursor, column=5, value=unit_cost)  # Unit. € in column E
                
                # Use formulas for totals to show calculation breakdown
                # Total € (one-off): =E{row}*C{row} (unit cost × quantity)
                ws.cell(row=row_cursor, column=6, value=f"=E{row_cursor}*C{row_cursor}")
                
                ws.cell(row=row_cursor, column=7, value=monthly_cost)  # Unit. € (monthly) in column G
                
                # Total € (monthly): =G{row}*C{row} (monthly unit cost × quantity)
                ws.cell(row=row_cursor, column=8, value=f"=G{row_cursor}*C{row_cursor}")
                
                # Yearly cost Total € spans 2 columns (I-J)
                # Total € (yearly): =H{row}*12 (monthly total × 12)
                ws.merge_cells(f'I{row_cursor}:J{row_cursor}')
                ws.cell(row=row_cursor, column=9, value=f"=H{row_cursor}*12")
                
                # Style the material row
                for col in range(2, 11):  # Columns B to J
                    cell = ws.cell(row=row_cursor, column=col)
                    cell.border = self.thin_border
                    
                    if col in [5, 6, 7, 8, 9]:  # Cost columns (now only 5 columns)
                        cell.number_format = self.currency_format_eur
                
                row_cursor += 1

            # Add SPACING after each category (empty row)
            if materials.exists():
                ws.append(["", "", "", "", "", "", "", "", "", ""]) # 10 empty cells for 10 columns
                row_cursor += 1

        # Update totals with formulas after all materials have been added
        self._update_totals_with_formulas(ws, row_cursor)

        # Auto-adjust column widths
        column_widths = {
            'A': 20,  # Categories (smaller)
            'B': 40,  # Item
            'C': 12,  # Quantity
            'D': 60,  # Description (BIGGER)
            'E': 12,  # Unit. €
            'F': 12,  # Total €
            'G': 12,  # Unit. € (monthly)
            'H': 12,  # Total € (monthly)
            'I': 12,  # Total € (yearly) - spans I-J
            'J': 12   # Total € (yearly) - spans I-J
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _update_totals_with_formulas(self, ws, last_row):
        """Update the summary totals with Excel formulas that show calculation breakdown"""
        # Find the range of material data rows (from row 23 to last_row)
        # Material data starts at row 23 and goes to last_row
        start_row = 23
        end_row = last_row - 1  # Subtract 1 to avoid empty rows at the end
        
        # Create formulas for each total - store only the formula, not text
        # One off total: SUM of column F (Total € for one-off costs)
        one_off_formula = f"=SUM(F{start_row}:F{end_row})"
        ws['F4'] = one_off_formula
        
        # Monthly total: SUM of column H (Total € for monthly costs)
        monthly_formula = f"=SUM(H{start_row}:H{end_row})"
        ws['F5'] = monthly_formula
        
        # Yearly total: SUM of column I (Total € for yearly costs)
        yearly_formula = f"=SUM(I{start_row}:I{end_row})"
        ws['F6'] = yearly_formula
        
        # MONTHLY TOTAL: One off + Monthly
        monthly_total_formula = f"=F4+F5"
        ws['F7'] = monthly_total_formula
        
        # YEARLY TOTAL: One off + Yearly  
        yearly_total_formula = f"=F4+F6"
        ws['F8'] = yearly_total_formula
        
        # Apply currency formatting to the formula cells
        for cell_ref in ['F4', 'F5', 'F6', 'F7', 'F8']:
            ws[cell_ref].number_format = self.currency_format_eur
    
    def _is_internet_service(self, material_name):
        """Check if a material is an internet service"""
        internet_types = ['Fiber Optic', 'STARLINK', 'VSAT']
        return any(internet_type in material_name for internet_type in internet_types)